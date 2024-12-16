from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Create a workbook and active worksheet
wb = Workbook()
ws = wb.active

# Define the headers and data
data = [
    ("ds1", "Entry 1"),
    ("ds1", "Entry 2"),
    ("ds2", "Entry 3"),
    ("ds3", "Entry 4"),
    ("ds4", "Entry 5"),
    ("ds4", "Entry 6"),
    ("ds4", "Entry 7"),
    ("ds5", "Entry 8"),
]

# Write headers
headers = ["Column 1", "Column 2"]
ws.append(headers)

# Write the data to the sheet
current_ds = None
start_row = None
for row_idx, (ds, entry) in enumerate(data, start=2):  # Start from row 2 (after headers)
    ws[f"A{row_idx}"] = ds
    ws[f"B{row_idx}"] = entry

    # Merge cells for repeated `ds`
    if ds == current_ds:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=row_idx, end_column=1)
    else:
        current_ds = ds
        start_row = row_idx

# Apply alignment to the merged cells
for row in range(2, len(data) + 2):
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top")

# Save the workbook
wb.save("merged_table.xlsx")

print("Excel file 'merged_table.xlsx' created with merged rows for ds1 entries.")
